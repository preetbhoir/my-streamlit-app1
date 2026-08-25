import io
from pathlib import Path
import base64
import re
import sqlite3
from datetime import datetime

import openpyxl
import pandas as pd
import plotly.express as px
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB_FILE = "school_assets.db"
BILLS_DIR = Path(__file__).resolve().parent / "asset_bills"
BILLS_DIR.mkdir(parents=True, exist_ok=True)

st.set_page_config(
    page_title="School Asset Manager",
    page_icon="🏫",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -----------------------------
# DATABASE
# -----------------------------
def get_db_connection():
    return sqlite3.connect(DB_FILE)

def init_db():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS assets (
                asset_code TEXT PRIMARY KEY,
                description TEXT,
                category TEXT,
                asset_group TEXT,
                location TEXT,
                department TEXT,
                asset_user TEXT,
                purchase_date TEXT,
                cost REAL,
                expiry_date TEXT,
                vendor TEXT,
                status TEXT DEFAULT 'Active',
                bill_file TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS asset_movements (
                movement_id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_code TEXT,
                from_location TEXT,
                to_location TEXT,
                from_department TEXT,
                to_department TEXT,
                move_date TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS disposals (
                disposal_id INTEGER PRIMARY KEY AUTOINCREMENT,
                asset_code TEXT UNIQUE,
                disposal_date TEXT,
                disposal_cost REAL,
                reason TEXT
            )
        """)
        conn.commit()

init_db()

# BILL_COLUMN_MIGRATION: keep older databases compatible.
try:
    with get_db_connection() as conn:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(assets)").fetchall()]
        if "bill_file" not in cols:
            conn.execute("ALTER TABLE assets ADD COLUMN bill_file TEXT")
            conn.commit()
except Exception:
    pass

# -----------------------------
# HIGH-END UI
# -----------------------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

.stApp {
    background: #f5f7fb;
}

section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #111827 0%, #172554 100%);
    border-right: 1px solid #263451;
}

section[data-testid="stSidebar"] * {
    color: #e5e7eb !important;
}

section[data-testid="stSidebar"] .stRadio label {
    padding: 10px 12px;
    border-radius: 10px;
    margin: 2px 0;
}

.hero {
    padding: 26px 30px;
    border-radius: 20px;
    background: linear-gradient(135deg, #0f172a 0%, #1d4ed8 55%, #2563eb 100%);
    color: white;
    margin-bottom: 22px;
    box-shadow: 0 14px 35px rgba(15, 23, 42, .14);
}

.hero h1 {
    margin: 0;
    font-size: 32px;
    font-weight: 800;
}

.hero p {
    margin: 7px 0 0;
    color: #dbeafe;
    font-size: 14px;
}

.card {
    background: white;
    border: 1px solid #e5e7eb;
    border-radius: 16px;
    padding: 18px 20px;
    box-shadow: 0 7px 20px rgba(15, 23, 42, .06);
}

.card-title {
    color: #64748b;
    font-size: 12px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: .06em;
}

.card-value {
    color: #0f172a;
    font-size: 29px;
    font-weight: 800;
    margin-top: 5px;
}

.section-title {
    font-size: 20px;
    font-weight: 750;
    color: #0f172a;
    margin: 10px 0 14px;
}

.small-note {
    color: #64748b;
    font-size: 12px;
}

div[data-testid="stDataFrame"] {
    border-radius: 14px;
    overflow: hidden;
}

button[kind="primary"] {
    border-radius: 10px;
    font-weight: 700;
}

.stButton > button {
    border-radius: 10px;
    font-weight: 600;
}

div[data-testid="stMetric"] {
    background: white;
    border: 1px solid #e5e7eb;
    padding: 15px;
    border-radius: 14px;
    box-shadow: 0 6px 18px rgba(15, 23, 42, .05);
}

div[data-baseweb="input"] > div,
div[data-baseweb="select"] > div,
textarea {
    border-radius: 10px !important;
}

hr {
    border-color: #e5e7eb;
}
</style>
""", unsafe_allow_html=True)

# -----------------------------
# HELPERS
# -----------------------------
def metric_card(title, value, subtitle=""):
    st.markdown(
        f"""
        <div class="card">
            <div class="card-title">{title}</div>
            <div class="card-value">{value}</div>
            <div class="small-note">{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

def load_assets():
    with get_db_connection() as conn:
        return pd.read_sql("""
            SELECT a.asset_code, a.description, a.category, a.asset_group,
                   a.location, a.department, a.asset_user, a.vendor,
                   a.purchase_date, a.cost, a.expiry_date, a.status,
                   COALESCE(a.bill_file, '') AS bill_file,
                   CASE WHEN a.status='Discarded'
                        THEN COALESCE(d.reason, 'Discarded')
                        ELSE 'Active'
                   END AS remark
            FROM assets a
            LEFT JOIN disposals d ON a.asset_code=d.asset_code
        """, conn)

def load_movements():
    with get_db_connection() as conn:
        return pd.read_sql("SELECT * FROM asset_movements ORDER BY movement_id DESC", conn)

def load_disposals():
    with get_db_connection() as conn:
        return pd.read_sql("SELECT * FROM disposals ORDER BY disposal_id DESC", conn)

def generate_excel_export(selected_tabs):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="000000")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    data_map = {
        "Asset Inventory": load_assets(),
        "Movement History": load_movements(),
        "Disposals": load_disposals(),
    }

    for name in selected_tabs:
        df = data_map[name]
        ws = wb.create_sheet(title=name)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(max_len + 3, 45))

    wb.save(output)
    output.seek(0)
    return output

# -----------------------------
# SIDEBAR
# -----------------------------
st.sidebar.markdown("## 🏫 Asset Manager")
st.sidebar.caption("School IT & Asset Control")
st.sidebar.divider()

menu = st.sidebar.radio(
    "MAIN MENU",
    [
        "📊 Dashboard",
        "📋 Asset Inventory",
        "➕ Add Asset",
        "📁 Import Excel",
        "🚚 Transfer Asset",
        "⏰ Warranty Alerts",
        "🗑️ Disposal",
        "📤 Reports",
    ],
)

st.sidebar.divider()
st.sidebar.markdown("**System**")
st.sidebar.caption(f"Database: `{DB_FILE}`")
st.sidebar.caption(f"Last opened: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")

df = load_assets()

# -----------------------------
# DASHBOARD
# -----------------------------
if menu == "📊 Dashboard":
    st.markdown("""
    <div class="hero">
        <h1>🏫 School Asset Management</h1>
        <p>Central dashboard for inventory, transfers, warranty tracking, disposal and reporting.</p>
    </div>
    """, unsafe_allow_html=True)

    active = df[df.status == "Active"]
    discarded = df[df.status == "Discarded"]

    today = pd.Timestamp(datetime.now().date())
    exp = active.copy()
    exp["expiry_dt"] = pd.to_datetime(exp["expiry_date"], format="mixed", dayfirst=True, errors="coerce")
    expired = int((exp["expiry_dt"] < today).sum())
    soon = int(((exp["expiry_dt"] >= today) & (exp["expiry_dt"] <= today + pd.Timedelta(days=90))).sum())
    total_value = float(active["cost"].fillna(0).sum())

    c1, c2, c3, c4 = st.columns(4)
    with c1: metric_card("Active Assets", f"{len(active):,}", "Currently in service")
    with c2: metric_card("Discarded", f"{len(discarded):,}", "Disposed / scrapped")
    with c3: metric_card("Warranty Alerts", f"{expired + soon:,}", f"{expired} expired • {soon} due within 90 days")
    with c4: metric_card("Asset Value", f"₹{total_value:,.0f}", "Active asset purchase value")

    st.markdown("<div class='section-title'>📈 Asset Overview</div>", unsafe_allow_html=True)

    left, right = st.columns([1.15, 1])
    with left:
        if not active.empty:
            cat = active["category"].fillna("Uncategorized").value_counts().reset_index()
            cat.columns = ["Category", "Count"]
            fig = px.bar(cat, x="Category", y="Count", title="Assets by Category", text_auto=True)
            fig.update_layout(template="plotly_white", height=360, margin=dict(l=10,r=10,t=55,b=10))
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No active assets available.")

    with right:
        if not active.empty:
            loc = active["location"].fillna("Unassigned").value_counts().reset_index()
            loc.columns = ["Location", "Count"]
            fig = px.pie(loc, names="Location", values="Count", hole=.55, title="Location Distribution")
            fig.update_layout(template="plotly_white", height=360, margin=dict(l=10,r=10,t=55,b=10))
            st.plotly_chart(fig, use_container_width=True)

    st.markdown("<div class='section-title'>⚠️ Warranty Watch</div>", unsafe_allow_html=True)
    if not exp.empty:
        alert = exp[(exp["expiry_dt"].notna()) & (exp["expiry_dt"] <= today + pd.Timedelta(days=90))].copy()
        if not alert.empty:
            alert["expiry_date"] = alert["expiry_dt"].dt.strftime("%d-%m-%Y")
            st.dataframe(
                alert[["asset_code","description","category","location","vendor","expiry_date"]],
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.success("No assets require warranty attention in the next 90 days.")

# -----------------------------
# INVENTORY
# -----------------------------
elif menu == "📋 Asset Inventory":
    st.markdown("<div class='section-title'>📋 Asset Inventory</div>", unsafe_allow_html=True)

    if df.empty:
        st.info("No assets found in the database.")
    else:
        f1, f2, f3, f4 = st.columns(4)
        categories = ["All"] + sorted(df["category"].dropna().astype(str).unique().tolist())
        locations = ["All"] + sorted(df["location"].dropna().astype(str).unique().tolist())
        departments = ["All"] + sorted(df["department"].dropna().astype(str).unique().tolist())
        statuses = ["All"] + sorted(df["status"].dropna().astype(str).unique().tolist())

        cat = f1.selectbox("Category", categories)
        loc = f2.selectbox("Location", locations)
        dept = f3.selectbox("Department", departments)
        status = f4.selectbox("Status", statuses)
        search = st.text_input("🔎 Search asset code, description, vendor, user...")

        filtered = df.copy()
        if cat != "All": filtered = filtered[filtered.category == cat]
        if loc != "All": filtered = filtered[filtered.location == loc]
        if dept != "All": filtered = filtered[filtered.department == dept]
        if status != "All": filtered = filtered[filtered.status == status]
        if search:
            filtered = filtered[filtered.astype(str).apply(
                lambda col: col.str.contains(search, case=False, na=False)
            ).any(axis=1)]

        a, b, c = st.columns(3)
        with a: metric_card("Filtered Records", f"{len(filtered):,}")
        with b: metric_card("Active", f"{(filtered.status == 'Active').sum():,}")
        with c: metric_card("Total Cost", f"₹{filtered.cost.fillna(0).sum():,.0f}")

        st.dataframe(filtered, use_container_width=True, hide_index=True)


        st.markdown("### 📎 Purchase Bill / Invoice")
        bill_rows = filtered[filtered["bill_file"].fillna("").astype(str).str.strip() != ""].copy()
        if bill_rows.empty:
            st.info("No scanned bill is attached to the assets shown by the current filters.")
        else:
            bill_options = {f"{r.asset_code} — {r.description}": r.asset_code for _, r in bill_rows.iterrows()}
            selected_bill_asset = st.selectbox("Select asset to view its bill", list(bill_options.keys()), key="attached_bill_asset")
            selected_code = bill_options[selected_bill_asset]
            selected_row = bill_rows[bill_rows.asset_code == selected_code].iloc[0]
            bill_name = str(selected_row.get("bill_file", "")).strip()
            bill_path = BILLS_DIR / bill_name
            if bill_name and bill_path.exists():
                st.caption(f"Attached file: {bill_name}")
                suffix = bill_path.suffix.lower()
                if suffix == ".pdf":
                    pdf_b64 = base64.b64encode(bill_path.read_bytes()).decode("utf-8")
                    html = f"<iframe src='data:application/pdf;base64,{pdf_b64}' width='100%' height='700' style='border:1px solid #d1d5db;border-radius:12px;'></iframe>"
                    st.components.v1.html(html, height=720)
                    mime = "application/pdf"
                else:
                    st.image(str(bill_path), caption=f"Bill — {selected_code}", use_container_width=True)
                    mime = "image/png" if suffix == ".png" else "image/jpeg"
                st.download_button("⬇️ Download Attached Bill", data=bill_path.read_bytes(), file_name=bill_path.name, mime=mime, key=f"download_bill_{selected_code}")
            else:
                st.error(f"The database has a bill name ({bill_name}), but the file was not found in: {BILLS_DIR}")

# -----------------------------
# ADD ASSET
# -----------------------------
elif menu == "➕ Add Asset":
    st.markdown("<div class='section-title'>➕ Register New Asset</div>", unsafe_allow_html=True)
    st.caption("Enter the asset details below. Asset Code and Description are mandatory.")

    with st.form("add_asset", clear_on_submit=True):
        a, b, c = st.columns(3)
        with a:
            code = st.text_input("Asset Code *", placeholder="GGHS/CAM/101")
            description = st.text_input("Description *", placeholder="CCTV Camera")
            category = st.text_input("Category", placeholder="Camera / Computer / Furniture")
            group = st.text_input("Asset Group")
        with b:
            location = st.text_input("Location / Floor")
            department = st.text_input("Department")
            user = st.text_input("Asset User / Custodian")
            vendor = st.text_input("Vendor")
        with c:
            purchase = st.date_input("Purchase Date", value=datetime.now())
            cost = st.number_input("Cost (₹)", min_value=0.0, step=100.0)
            expiry = st.date_input("Warranty / Life Expiry", value=datetime.now())

        st.markdown("### 📎 Purchase Bill / Invoice")
        bill_upload = st.file_uploader("Upload scanned bill copy", type=["pdf", "jpg", "jpeg", "png"], help="Attach the bill to this asset.")
        save = st.form_submit_button("💾 Save Asset", type="primary", use_container_width=True)

    if save:
        if not code.strip() or not description.strip():
            st.error("Asset Code and Description are required.")
        else:
            saved_bill = None
            try:
                with get_db_connection() as conn:
                    conn.execute("""
                        INSERT INTO assets
                        (asset_code, description, category, asset_group, location, department,
                         asset_user, purchase_date, cost, expiry_date, vendor, status, bill_file)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active', ?)
                    """, (code.strip(), description.strip(), category.strip(), group.strip(), location.strip(), department.strip(), user.strip(), purchase.isoformat(), cost, expiry.isoformat(), vendor.strip(), None))
                    conn.commit()
                    if bill_upload is not None:
                        safe_code = re.sub(r"[^A-Za-z0-9_-]", "_", code.strip())
                        ext = Path(bill_upload.name).suffix.lower()
                        saved_bill = f"{safe_code}_bill_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
                        bill_path = BILLS_DIR / saved_bill
                        bill_path.write_bytes(bill_upload.getbuffer())
                        conn.execute("UPDATE assets SET bill_file=? WHERE asset_code=?", (saved_bill, code.strip()))
                        conn.commit()
                st.success(f"Asset {code} added successfully" + (" with bill attached." if saved_bill else "."))
                st.rerun()
            except sqlite3.IntegrityError:
                if saved_bill: (BILLS_DIR / saved_bill).unlink(missing_ok=True)
                st.error(f"Asset Code '{code}' already exists.")
            except Exception as ex:
                if saved_bill: (BILLS_DIR / saved_bill).unlink(missing_ok=True)
                st.error(f"Could not save asset or bill: {ex}")

# -----------------------------
# IMPORT
# -----------------------------

    st.markdown("### 📎 Scan Bill / Purchase Invoice")
    bill_upload = st.file_uploader(
        "Upload scanned bill copy",
        type=["pdf", "jpg", "jpeg", "png"],
        help="Attach the purchase bill/invoice to this asset. PDF or image files are supported."
    )

elif menu == "📁 Import Excel":
    st.markdown("<div class='section-title'>📁 Bulk Import from Excel</div>", unsafe_allow_html=True)
    st.info("Upload an .xlsx file containing a sheet named **Assets**.")

    uploaded = st.file_uploader("Choose Excel file", type=["xlsx"])
    if uploaded:
        try:
            upload_df = pd.read_excel(uploaded, sheet_name="Assets")
            st.success(f"Loaded {len(upload_df):,} rows.")
            st.dataframe(upload_df.head(10), use_container_width=True, hide_index=True)

            if st.button("⬆️ Import All Assets", type="primary", use_container_width=True):
                count = 0
                with get_db_connection() as conn:
                    for _, row in upload_df.iterrows():
                        code = str(row.get("Asset Code", "")).strip()
                        if not code or code == "nan":
                            continue
                        p = pd.to_datetime(row.get("Purchase Date(dd/mm/yyyy)", ""), errors="coerce")
                        e = pd.to_datetime(row.get("Expiry Date(dd/mm/yyyy)", ""), errors="coerce")
                        conn.execute("""
                            INSERT OR REPLACE INTO assets
                            (asset_code, description, category, asset_group, location, department,
                             asset_user, purchase_date, cost, expiry_date, vendor, status)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active')
                        """, (
                            code,
                            str(row.get("Asset Description", "")).strip(),
                            str(row.get("Category", "")).strip(),
                            str(row.get("Asset Group", "")).strip(),
                            str(row.get("Asset Location", "")).strip(),
                            str(row.get("Department", "")).strip(),
                            str(row.get("Asset User", "")).strip(),
                            p.strftime("%Y-%m-%d") if pd.notnull(p) else "",
                            float(row.get("Cost", 0) or 0),
                            e.strftime("%Y-%m-%d") if pd.notnull(e) else "",
                            str(row.get("Vendor", "")).strip(),
                        ))
                        count += 1
                    conn.commit()
                st.success(f"Successfully imported {count:,} assets.")
                st.rerun()
        except Exception as e:
            st.error(f"Could not process the Excel file: {e}")

# -----------------------------
# TRANSFER
# -----------------------------
elif menu == "🚚 Transfer Asset":
    st.markdown("<div class='section-title'>🚚 Transfer Asset</div>", unsafe_allow_html=True)
    active = df[df.status == "Active"].copy()

    if active.empty:
        st.info("No active assets available for transfer.")
    else:
        f1, f2 = st.columns(2)
        cats = ["All"] + sorted(active.category.dropna().astype(str).unique().tolist())
        depts = ["All"] + sorted(active.department.dropna().astype(str).unique().tolist())
        fc = f1.selectbox("Filter Category", cats)
        fd = f2.selectbox("Filter Department", depts)

        work = active.copy()
        if fc != "All": work = work[work.category == fc]
        if fd != "All": work = work[work.department == fd]

        if not work.empty:
            options = {
                f"{r.asset_code} — {r.description} ({r.location})": r.asset_code
                for _, r in work.iterrows()
            }
            selected = st.selectbox("Select Asset", list(options.keys()))
            code = options[selected]
            current = work[work.asset_code == code].iloc[0]

            st.markdown(
                f"<div class='card'><b>Current Location:</b> {current.location} &nbsp; "
                f"<b>Department:</b> {current.department}</div>",
                unsafe_allow_html=True,
            )
            st.write("")
            n1, n2 = st.columns(2)
            new_loc = n1.text_input("New Location")
            new_dept = n2.text_input("New Department")

            if st.button("🔄 Confirm Transfer", type="primary", use_container_width=True):
                if not new_loc.strip() or not new_dept.strip():
                    st.error("New Location and New Department are required.")
                else:
                    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    with get_db_connection() as conn:
                        conn.execute("""
                            INSERT INTO asset_movements
                            (asset_code, from_location, to_location, from_department, to_department, move_date)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (code, current.location, new_loc, current.department, new_dept, now))
                        conn.execute(
                            "UPDATE assets SET location=?, department=? WHERE asset_code=?",
                            (new_loc, new_dept, code)
                        )
                        conn.commit()
                    st.success(f"{code} transferred successfully.")
                    st.rerun()
        else:
            st.warning("No assets match the selected filters.")

# -----------------------------
# WARRANTY
# -----------------------------
elif menu == "⏰ Warranty Alerts":
    st.markdown("<div class='section-title'>⏰ Warranty & Expiry Center</div>", unsafe_allow_html=True)
    exp = df[df.status == "Active"].copy()
    exp["expiry_dt"] = pd.to_datetime(exp["expiry_date"], format="mixed", dayfirst=True, errors="coerce")
    exp = exp.dropna(subset=["expiry_dt"])
    today = pd.Timestamp(datetime.now().date())
    soon = today + pd.Timedelta(days=90)

    expired = exp[exp.expiry_dt < today]
    soon_df = exp[(exp.expiry_dt >= today) & (exp.expiry_dt <= soon)]
    valid = exp[exp.expiry_dt > soon]

    c1, c2, c3 = st.columns(3)
    with c1: metric_card("🔴 Expired", len(expired), "Warranty already ended")
    with c2: metric_card("🟡 Due Soon", len(soon_df), "Within next 90 days")
    with c3: metric_card("🟢 Valid", len(valid), "More than 90 days remaining")

    s1, s2 = st.columns(2)
    categories = ["All"] + sorted(exp.category.dropna().astype(str).unique().tolist())
    selected_cat = s1.selectbox("Category", categories)
    selected_status = s2.selectbox("Expiry Status", ["All", "Expired", "Due Soon", "Valid"])

    work = exp.copy()
    if selected_cat != "All":
        work = work[work.category == selected_cat]
    if selected_status == "Expired":
        work = work[work.expiry_dt < today]
    elif selected_status == "Due Soon":
        work = work[(work.expiry_dt >= today) & (work.expiry_dt <= soon)]
    elif selected_status == "Valid":
        work = work[work.expiry_dt > soon]

    work["expiry_date"] = work.expiry_dt.dt.strftime("%d-%m-%Y")
    st.dataframe(
        work[["asset_code","description","category","location","vendor","expiry_date"]],
        use_container_width=True,
        hide_index=True,
    )

# -----------------------------
# DISPOSAL
# -----------------------------
elif menu == "🗑️ Disposal":
    st.markdown("<div class='section-title'>🗑️ Asset Disposal</div>", unsafe_allow_html=True)
    active = df[df.status == "Active"].copy()

    if active.empty:
        st.info("No active assets available for disposal.")
    else:
        f1, f2 = st.columns(2)
        cats = ["All"] + sorted(active.category.dropna().astype(str).unique().tolist())
        locs = ["All"] + sorted(active.location.dropna().astype(str).unique().tolist())
        fc = f1.selectbox("Category", cats)
        fl = f2.selectbox("Location", locs)

        work = active.copy()
        if fc != "All": work = work[work.category == fc]
        if fl != "All": work = work[work.location == fl]

        if not work.empty:
            options = {
                f"{r.asset_code} — {r.description} ({r.location})": r.asset_code
                for _, r in work.iterrows()
            }
            selected = st.selectbox("Select Asset to Discard", list(options.keys()))
            code = options[selected]

            scrap = st.number_input("Scrap Value (₹)", min_value=0.0, step=100.0)
            reason = st.text_area("Reason for Disposal", placeholder="Damaged, obsolete, beyond repair...")

            if st.button("⚠️ Mark as Discarded", type="primary", use_container_width=True):
                with get_db_connection() as conn:
                    conn.execute(
                        "INSERT INTO disposals (asset_code, disposal_date, disposal_cost, reason) VALUES (?, ?, ?, ?)",
                        (code, datetime.now().strftime("%Y-%m-%d"), scrap, reason.strip())
                    )
                    conn.execute("UPDATE assets SET status='Discarded' WHERE asset_code=?", (code,))
                    conn.commit()
                st.success(f"Asset {code} has been marked as discarded.")
                st.rerun()
        else:
            st.warning("No assets match the selected filters.")

# -----------------------------
# REPORTS
# -----------------------------
elif menu == "📤 Reports":
    st.markdown("<div class='section-title'>📤 Asset Reports</div>", unsafe_allow_html=True)
    st.caption("Use the same filters as Asset Inventory, preview the filtered records, and export a professional Excel or PDF report.")

    # Same Asset Inventory filters
    report_df = df.copy()
    f1, f2, f3, f4 = st.columns(4)

    categories = ["All"] + sorted(report_df["category"].dropna().astype(str).unique().tolist())
    locations = ["All"] + sorted(report_df["location"].dropna().astype(str).unique().tolist())
    departments = ["All"] + sorted(report_df["department"].dropna().astype(str).unique().tolist())
    statuses = ["All"] + sorted(report_df["status"].dropna().astype(str).unique().tolist())

    report_cat = f1.selectbox("Category", categories, key="report_cat")
    report_loc = f2.selectbox("Location", locations, key="report_loc")
    report_dept = f3.selectbox("Department", departments, key="report_dept")
    report_status = f4.selectbox("Status", statuses, key="report_status")
    report_search = st.text_input("🔎 Keyword Search", key="report_search")

    if report_cat != "All":
        report_df = report_df[report_df["category"].astype(str) == report_cat]
    if report_loc != "All":
        report_df = report_df[report_df["location"].astype(str) == report_loc]
    if report_dept != "All":
        report_df = report_df[report_df["department"].astype(str) == report_dept]
    if report_status != "All":
        report_df = report_df[report_df["status"].astype(str) == report_status]
    if report_search.strip():
        term = report_search.strip().lower()
        report_df = report_df[
            report_df.astype(str).apply(
                lambda c: c.str.lower().str.contains(term, na=False)
            ).any(axis=1)
        ]

    total_cost = pd.to_numeric(report_df["cost"], errors="coerce").fillna(0).sum()

    c1, c2, c3 = st.columns(3)
    with c1:
        metric_card("Filtered Records", f"{len(report_df):,}", "Assets in report")
    with c2:
        metric_card("Total Cost", f"₹{total_cost:,.2f}", "Filtered purchase value")
    with c3:
        metric_card("Active Assets", f"{(report_df['status'] == 'Active').sum():,}", "Currently active")

    st.markdown("### 👁️ Report Preview")
    st.dataframe(report_df, use_container_width=True, hide_index=True)

    # Clean report columns
    rename_map = {
        "asset_code": "Asset Code",
        "description": "Description",
        "category": "Category",
        "asset_group": "Asset Group",
        "location": "Location / Floor",
        "department": "Department",
        "asset_user": "Asset User / Custodian",
        "vendor": "Vendor",
        "purchase_date": "Purchase Date",
        "cost": "Cost (₹)",
        "expiry_date": "Warranty / Expiry",
        "status": "Status",
        "remark": "Remark",
    }

    cols = [c for c in rename_map if c in report_df.columns]
    export_df = report_df[cols].copy().rename(columns=rename_map)

    for col in ["Purchase Date", "Warranty / Expiry"]:
        if col in export_df.columns:
            dates = pd.to_datetime(export_df[col], format="mixed", dayfirst=True, errors="coerce")
            export_df[col] = dates.dt.strftime("%d-%m-%Y").fillna("")

    if "Cost (₹)" in export_df.columns:
        export_df["Cost (₹)"] = pd.to_numeric(
            export_df["Cost (₹)"], errors="coerce"
        ).fillna(0)

    st.divider()
    st.markdown("### 📄 Export")

    report_title = st.text_input(
        "Report Title",
        value="School Asset Inventory Report",
        key="report_title"
    )

    ex1, ex2 = st.columns(2)

    # ---------------- EXCEL ----------------
    with ex1:
        excel = io.BytesIO()

        with pd.ExcelWriter(excel, engine="openpyxl") as writer:
            export_df.to_excel(
                writer,
                index=False,
                sheet_name="Asset Inventory",
                startrow=5
            )

            ws = writer.book["Asset Inventory"]
            last_col = max(1, len(export_df.columns))

            ws["A1"] = report_title
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
            ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
            ws["A1"].fill = PatternFill("solid", fgColor="17365D")
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 30

            ws["A2"] = "Generated"
            ws["B2"] = datetime.now().strftime("%d-%m-%Y %I:%M %p")
            ws["C2"] = "Category"
            ws["D2"] = report_cat
            ws["E2"] = "Location"
            ws["F2"] = report_loc

            ws["A3"] = "Records"
            ws["B3"] = len(export_df)
            ws["C3"] = "Department"
            ws["D3"] = report_dept
            ws["E3"] = "Status"
            ws["F3"] = report_status

            header_row = 6
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF")
            thin = Side(style="thin", color="D9D9D9")

            for cell in ws[header_row]:
                if cell.value is not None:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
                for cell in row:
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            ws.freeze_panes = "A7"
            ws.auto_filter.ref = f"A6:{get_column_letter(ws.max_column)}{ws.max_row}"

            for col in ws.columns:
                values = [len(str(c.value or "")) for c in col]
                width = min(max(max(values) + 3, 12), 35)
                ws.column_dimensions[get_column_letter(col[0].column)].width = width

            # Total row for the filtered assets
            total_row = ws.max_row + 2
            ws.cell(total_row, 1).value = "TOTAL ASSETS"
            ws.cell(total_row, 2).value = len(export_df)

            for i, cell in enumerate(ws[header_row], start=1):
                if cell.value == "Cost (₹)":
                    for r in range(7, total_row):
                        ws.cell(r, i).number_format = '₹#,##0.00'
                    ws.cell(total_row, i).value = total_cost
                    ws.cell(total_row, i).number_format = '₹#,##0.00'

            total_fill = PatternFill("solid", fgColor="D9EAF7")
            total_font = Font(bold=True, color="17365D")
            for cell in ws[total_row]:
                cell.fill = total_fill
                cell.font = total_font
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        excel.seek(0)

        st.download_button(
            "📊 Download Excel Report",
            data=excel.getvalue(),
            file_name=f"School_Asset_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    # ---------------- PDF ----------------
    with ex2:
        pdf = io.BytesIO()

        doc = SimpleDocTemplate(
            pdf,
            pagesize=landscape(A4),
            leftMargin=20,
            rightMargin=20,
            topMargin=20,
            bottomMargin=20,
        )

        styles = getSampleStyleSheet()
        story = [
            Paragraph(report_title, styles["Title"]),
            Spacer(1, 6),
            Paragraph(
                f"Generated: {datetime.now().strftime('%d-%m-%Y %I:%M %p')} | "
                f"Records: {len(export_df):,} | Total Cost: ₹{total_cost:,.2f}",
                styles["Normal"]
            ),
            Spacer(1, 5),
            Paragraph(
                f"Filters: Category={report_cat} | Location={report_loc} | "
                f"Department={report_dept} | Status={report_status}",
                styles["Normal"]
            ),
            Spacer(1, 10),
        ]

        pdf_df = export_df.fillna("").copy()

        if "Cost (₹)" in pdf_df.columns:
            pdf_df["Cost (₹)"] = pdf_df["Cost (₹)"].apply(
                lambda x: f"₹{float(x):,.2f}" if str(x) else ""
            )

        # Shorter PDF headings so the report fits landscape A4.
        pdf_df.rename(
            columns={
                "Location / Floor": "Location",
                "Asset User / Custodian": "Custodian",
                "Warranty / Expiry": "Expiry",
            },
            inplace=True
        )

        table_data = [list(pdf_df.columns)]
        table_data += pdf_df.astype(str).values.tolist()

        if len(table_data) == 1:
            table_data.append(["No records"] + [""] * (len(table_data[0]) - 1))

        # Total row for the filtered assets
        total_pdf_row = ["TOTAL ASSETS", str(len(export_df))]
        total_pdf_row += [""] * (len(table_data[0]) - len(total_pdf_row))
        table_data.append(total_pdf_row)

        page_width = landscape(A4)[0] - 40
        n = len(table_data[0])
        widths = [page_width / n] * n

        # Give text-heavy columns more room.
        for i, name in enumerate(table_data[0]):
            if name in ["Description", "Vendor", "Custodian"]:
                widths[i] *= 1.4

        scale = page_width / sum(widths)
        widths = [w * scale for w in widths]

        table = Table(table_data, colWidths=widths, repeatRows=1)

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D9D9D9")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9EAF7")),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#17365D")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor("#17365D")),
        ]))

        story.append(table)
        doc.build(story)
        pdf.seek(0)

        st.download_button(
            "📄 Download PDF Report",
            data=pdf.getvalue(),
            file_name=f"School_Asset_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
            mime="application/pdf",
            use_container_width=True,
            type="primary",
        )

